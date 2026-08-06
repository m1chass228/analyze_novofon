import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
import json
import os
from datetime import datetime, timedelta

# Базовые пути для хранения отчетов
BASE_REPORTS_DIR = "reports"
INDIVIDUAL_DIR = os.path.join(BASE_REPORTS_DIR, "individual")
DAILY_DIR = os.path.join(BASE_REPORTS_DIR, "daily")
MASTER_REPORT_PATH = os.path.join(BASE_REPORTS_DIR, "master_calls_report.xlsx")


def init_report_structure():
    """Создает структуру папок для отчетов"""
    os.makedirs(INDIVIDUAL_DIR, exist_ok=True)
    os.makedirs(DAILY_DIR, exist_ok=True)


def create_call_report(gemini_json_str: str, call_id: str, duration_sec: int, record_url: str, customer_phone: str = "Не определен", call_start_time: str = None, direction: str = "in", *args, **kwargs) -> str:
    """Создает премиальный индивидуальный отчет по одному звонку (Подробные баллы с комментариями + Резюме)"""
    try:
        data = json.loads(gemini_json_str)
    except Exception as e:
        print(f"Ошибка парсинга JSON: {e}")
        return ""

    init_report_structure()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Аудит Звонка {call_id}"
    
    ws.views.sheetView[0].showGridLines = True

    # ==================== ДИЗАЙНЕРСКАЯ ПАЛИТРА СТИЛЕЙ ====================
    font_main_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    font_section_title = Font(name="Segoe UI", size=12, bold=True, color="2C3E50")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="2C3E50")
    font_regular = Font(name="Segoe UI", size=10, color="2C3E50")
    font_section_header = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")
    
    fill_main_banner = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    fill_label_side = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    fill_cat_perfect = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
    fill_cat_good = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
    fill_cat_normal = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
    fill_cat_bad = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

    border_light = Border(left=Side(style="thin", color="D5D8DC"), right=Side(style="thin", color="D5D8DC"),
                          top=Side(style="thin", color="D5D8DC"), bottom=Side(style="thin", color="D5D8DC"))
    border_total = Border(top=Side(style="thin", color="2C3E50"), bottom=Side(style="double", color="2C3E50"))

    align_center = Alignment(horizontal="center", vertical="center")
    align_left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_label = Alignment(horizontal="left", vertical="center")

    # ==================== ЗАГОЛОВОК ДОКУМЕНТА ====================
    ws.merge_cells("A2:D2")
    title_cell = ws["A2"]
    title_cell.value = f"ОТЧЁТ ПО АУДИТУ КАЧЕСТВА РАЗГОВОРА № {call_id}"
    title_cell.font = font_main_title
    title_cell.fill = fill_main_banner
    title_cell.alignment = align_center
    ws.row_dimensions[2].height = 38

    score = int(data.get("total_score", 0))
    category = int(data.get("category", 4))

    if score >= 37: fill_result = fill_cat_perfect
    elif score >= 28: fill_result = fill_cat_good
    elif score >= 18: fill_result = fill_cat_normal
    else: fill_result = fill_cat_bad

    direction_readable = "Входящий" if direction == "in" else "Исходящий"

    display_date = "-"
    if call_start_time:
        try:
            dt = datetime.strptime(call_start_time.split(".")[0], "%Y-%m-%d %H:%M:%S")
            display_date = dt.strftime("%d.%m.%Y %H:%M")
        except:
            display_date = call_start_time
    else:
        display_date = datetime.now().strftime("%d.%m.%Y %H:%M")

    ai_name = data.get("admin_name", "Не определен")
    if ai_name.startswith("Не определен (") and ai_name.endswith(")"):
        ai_name = ai_name.replace("Не определен (", "").replace(")", "")

    # ==================== БЛОК МЕТАДАННЫХ ====================
    metadata = [
        ("ID звонка", call_id),
        ("Телефон клиента", str(customer_phone)),
        ("Направление звонка", direction_readable),
        ("Дата и время звонка", display_date),
        ("Длительность звонка", f"{duration_sec} сек (~{round(duration_sec/60, 1)} мин)"),
        ("Внутренний номер", ai_name),
        ("Итоговый балл", f"{score} / 43"),
        ("Категория качества", f"{category} категория")
    ]

    start_meta_row = 4
    for idx, (label, val) in enumerate(metadata):
        r = start_meta_row + idx
        ws.row_dimensions[r].height = 20
        
        c_label = ws.cell(row=r, column=1, value=label)
        c_label.font = font_bold
        c_label.fill = fill_label_side
        c_label.border = border_light
        c_label.alignment = align_label
        
        c_val = ws.cell(row=r, column=2, value=val)
        c_val.font = font_regular
        c_val.border = border_light
        c_val.alignment = align_label
        
        if label in ["Итоговый балл", "Категория качества"]:
            c_val.fill = fill_result
            c_val.font = font_bold

    r_link = start_meta_row + len(metadata)
    ws.row_dimensions[r_link].height = 22
    cl = ws.cell(row=r_link, column=1, value="Запись разговора")
    cl.font = font_bold
    cl.fill = fill_label_side
    cl.border = border_light
    
    cv = ws.cell(row=r_link, column=2)
    if record_url:
        cv.value = "▶ Слушать аудиозапись"
        cv.hyperlink = record_url
        cv.font = Font(name="Segoe UI", size=10, color="1A5276", underline="single", bold=True)
    else:
        cv.value = "Ссылка отсутствует"
        cv.font = font_regular
    cv.border = border_light

    # ==================== ПОДРОБНАЯ ТАБЛИЦА БАЛЛОВ С КОММЕНТАРИЯМИ ====================
    table_start_row = r_link + 2
    headers = ["Критерий чек-листа", "Балл", "Макс.", "Развернутый комментарий аналитика"]
    
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_main_banner
        cell.alignment = align_center
    ws.row_dimensions[table_start_row].height = 26

    criteria_map = {
        "contact": ("1. Установление контакта (до 2 б.)", 2),
        "name_usage": ("2. Использование имени (до 2 б.)", 2),
        "needs_analysis": ("3. Выяснение потребностей (до 9 б.)", 9),
        "presentation": ("4. Презентация услуг (до 7 б.)", 7),
        "pricing": ("5. Презентация цен (до 3 б.)", 3),
        "objections": ("6. Работа с возражениями (до 4 б.)", 4),
        "closing_to_appointment": ("7. Ведение к записи (до 8 б.)", 8),
        "termination": ("8. Завершение диалога (до 3 б.)", 3),
        "individual_approach": ("9. Индивидуальный подход (до 5 б.)", 5)
    }

    current_row = table_start_row + 1
    details = data.get("details", {})

    for key, (display_name, max_val) in criteria_map.items():
        ws.row_dimensions[current_row].height = 24
        item = details.get(key, {"score": 0, "comment": ""})
        score_val = item.get("score", 0)
        comment = item.get("comment", "-")

        c1 = ws.cell(row=current_row, column=1, value=display_name)
        c2 = ws.cell(row=current_row, column=2, value=score_val)
        c3 = ws.cell(row=current_row, column=3, value=max_val)
        c4 = ws.cell(row=current_row, column=4, value=comment)

        c1.font = font_bold
        c2.font = font_bold
        c2.alignment = align_center
        c3.font = font_regular
        c3.alignment = align_center
        c4.font = font_regular
        c4.alignment = align_left_wrap

        if score_val == 0:
            c2.fill = fill_cat_bad
        elif score_val == max_val:
            c2.fill = fill_cat_perfect

        for col in range(1, 5):
            ws.cell(row=current_row, column=col).border = border_light

        current_row += 1

    # Строка ИТОГО
    ws.row_dimensions[current_row].height = 26
    t1 = ws.cell(row=current_row, column=1, value="ИТОГО БАЛЛОВ:")
    t2 = ws.cell(row=current_row, column=2, value=f"=SUM(B{table_start_row+1}:B{current_row-1})")
    t3 = ws.cell(row=current_row, column=3, value=f"=SUM(C{table_start_row+1}:C{current_row-1})")
    ws.cell(row=current_row, column=4, value="")

    t1.font = font_bold
    t2.font = font_main_title
    t2.fill = fill_main_banner
    t2.alignment = align_center
    t3.font = font_bold
    t3.alignment = align_center

    for col in range(1, 5):
        ws.cell(row=current_row, column=col).border = border_total

    # ==================== БЛОК: КРАТКОЕ РЕЗЮМЕ ДИАЛОГА ====================
    r_overview_label = current_row + 2
    ws.merge_cells(f"A{r_overview_label}:D{r_overview_label}")
    o_lbl = ws.cell(row=r_overview_label, column=1, value="КРАТКОЕ РЕЗЮМЕ ДИАЛОГА (ОПИСАНИЕ):")
    o_lbl.font = font_section_title
    o_lbl.fill = font_section_header
    ws.row_dimensions[r_overview_label].height = 22

    r_overview_val = r_overview_label + 1
    ws.merge_cells(f"A{r_overview_val}:D{r_overview_val+3}")
    o_val = ws.cell(row=r_overview_val, column=1, value=data.get("dialog_overview", "Нет описания"))
    o_val.font = font_regular
    o_val.alignment = align_left_wrap
    
    ws.row_dimensions[r_overview_val].height = 18
    ws.row_dimensions[r_overview_val+1].height = 18
    ws.row_dimensions[r_overview_val+2].height = 18
    ws.row_dimensions[r_overview_val+3].height = 18

    # Проставляем границы для резюме
    for row_idx in range(r_overview_label, r_overview_val + 4):
        for col_idx in range(1, 5):
            ws.cell(row=row_idx, column=col_idx).border = border_light

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 75

    file_name = os.path.join(INDIVIDUAL_DIR, f"report_{call_id}.xlsx")
    wb.save(file_name)
    return file_name


def update_master_report(calls_data: list, public_urls_map: dict = None, *args, **kwargs) -> str:
    """Обновляет Сводный журнал (Строго пропускает звонки без резюме диалога)"""
    if public_urls_map is None:
        public_urls_map = {}

    init_report_structure()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводный журнал"
    
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A4"

    font_title = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    font_regular = Font(name="Segoe UI", size=9, color="2C3E50")
    font_bold = Font(name="Segoe UI", size=9, bold=True, color="2C3E50")
    font_link = Font(name="Segoe UI", size=9, color="1A5276", underline="single", bold=True)

    fill_main_banner = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    fill_zebra_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    fill_cat1 = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
    fill_cat2 = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
    fill_cat3 = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
    fill_cat4 = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

    border_light = Border(left=Side(style="thin", color="E5E7E9"), right=Side(style="thin", color="E5E7E9"),
                          top=Side(style="thin", color="E5E7E9"), bottom=Side(style="thin", color="E5E7E9"))

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "СВОДНЫЙ ЖУРНАЛ КОНТРОЛЯ КАЧЕСТВА И СТАТИСТИКИ ЗВОНКОВ"
    title_cell.font = font_title
    title_cell.fill = fill_main_banner
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 32

    headers = [
        "ID Звонка", "Дата", "Время", "Телефон клиента", "Длительность", "Итого балл (/43)", 
        "Внутр. номер", "Направление", "Запись", 
        "Резюме диалога", "Детальный отчёт", "Запись звонка"
    ]

    header_row = 3
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        cell.alignment = align_center
    ws.row_dimensions[header_row].height = 28

    current_row = header_row + 1

    # Индекс для зебры считаем отдельно, так как некоторые строки будут пропускаться
    added_rows_count = 0

    for item_tuple in calls_data:
        if isinstance(item_tuple, dict):
            start_time_str = item_tuple.get("start_time", "")
            call_id = item_tuple.get("call_id", "")
            duration = item_tuple.get("duration", 0)
            phone = item_tuple.get("customer_phone", "")
            admin_name_from_db = item_tuple.get("admin_name", "Не определен")
            analysis_text = item_tuple.get("analysis_text", "")
            record_url = item_tuple.get("record_url", "")
            direction = item_tuple.get("direction", "in")
            db_report_url = item_tuple.get("report_url", "")
        else:
            start_time_str = item_tuple[0]
            call_id = item_tuple[1]
            duration = item_tuple[2]
            phone = item_tuple[3]
            admin_name_from_db = item_tuple[4]
            analysis_text = item_tuple[6]
            record_url = item_tuple[7]
            direction = item_tuple[8] if len(item_tuple) > 8 else "in"
            db_report_url = item_tuple[9] if len(item_tuple) > 9 else ""

        total_score = 0
        appointment_made = False
        dialog_overview = ""

        if analysis_text:
            try:
                parsed = json.loads(analysis_text) if isinstance(analysis_text, str) else analysis_text
                total_score = int(parsed.get("total_score", 0))
                appointment_made = parsed.get("appointment_made", False)
                dialog_overview = parsed.get("dialog_overview", "")
            except:
                pass

        # === ЖЕСТКАЯ ПРОВЕРКА: Если резюме диалога пустое, полностью игнорируем этот звонок ===
        if not dialog_overview or str(dialog_overview).strip() in ["", "null", "None", "-"]:
            continue

        # Настраиваем высоту строки и зебру только для тех, кто прошел валидацию
        ws.row_dimensions[current_row].height = 28
        row_base_fill = fill_zebra_light if added_rows_count % 2 == 0 else None

        call_date = call_time = "-"
        if start_time_str:
            try:
                dt = datetime.strptime(start_time_str.split(".")[0], "%Y-%m-%d %H:%M:%S")
                call_date = dt.strftime("%d.%m.%Y")
                call_time = dt.strftime("%H:%M")
            except:
                pass

        direction_readable = "Входящий" if direction == "in" else "Исходящий"

        # Базовые данные (Колонки 1-9)
        row_data = [
            str(call_id), call_date, call_time, str(phone or "-"), int(duration or 0), total_score,
            str(admin_name_from_db or "Не определен"), direction_readable, "Да" if appointment_made else "Нет"
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = font_regular
            cell.border = border_light
            if row_base_fill: cell.fill = row_base_fill
            cell.alignment = align_center if col_idx not in [4, 7] else align_left

        # Подсветка "Итого балл" (колонка 6)
        score_cell = ws.cell(row=current_row, column=6)
        score_cell.font = font_bold
        if total_score >= 37: score_cell.fill = fill_cat1
        elif total_score >= 28: score_cell.fill = fill_cat2
        elif total_score >= 18: score_cell.fill = fill_cat3
        else: score_cell.fill = fill_cat4

        # Подсветка записи (колонка 9)
        booking_cell = ws.cell(row=current_row, column=9)
        if appointment_made:
            booking_cell.fill = fill_cat1
            booking_cell.font = font_bold

        # Описание диалога (Колонка J / 10)
        overview_cell = ws.cell(row=current_row, column=10, value=dialog_overview)
        overview_cell.font = font_regular
        overview_cell.border = border_light
        overview_cell.alignment = align_left
        if row_base_fill: overview_cell.fill = row_base_fill

        # Детальный отчёт по ссылке (Колонка K / 11)
        real_public_url = public_urls_map.get(str(call_id)) or db_report_url
        report_cell = ws.cell(row=current_row, column=11)
        
        if real_public_url and str(real_public_url).strip() not in ["", "None", "null", "-"]:
            report_cell.value = "Открыть отчёт"
            report_cell.hyperlink = real_public_url
            report_cell.font = font_link
        else:
            report_cell.value = "Не синхронизирован"
            report_cell.font = font_regular
            
        report_cell.alignment = align_center
        report_cell.border = border_light
        if row_base_fill: report_cell.fill = row_base_fill

        # Ссылка на запись звонка (Колонка L / 12)
        link_cell = ws.cell(row=current_row, column=12)
        if record_url:
            link_cell.value = "Слушать запись"
            link_cell.hyperlink = record_url
            link_cell.font = font_link
        else:
            link_cell.value = "-"
            link_cell.font = font_regular
        link_cell.alignment = align_center
        link_cell.border = border_light
        if row_base_fill: link_cell.fill = row_base_fill

        current_row += 1
        added_rows_count += 1

    # Ширины колонок
    column_widths = {
        'A': 14, 'B': 13, 'C': 10, 'D': 18, 'E': 14, 'F': 17, 'G': 28, 'H': 14, 'I': 11,
        'J': 45, 'K': 20, 'L': 18
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(MASTER_REPORT_PATH)
    return MASTER_REPORT_PATH

def update_daily_report(calls_data: list, public_urls_map: dict = None) -> str | None:
    """
    Создает или обновляет посуточный отчет в папке reports/daily/.
    Имя файла и вкладки: текущая дата (например, 2026-05-26.xlsx).
    Включает в себя только звонки за текущие сутки с жесткой проверкой резюме.
    """
    if public_urls_map is None:
        public_urls_map = {}

    # Проверяем, что папки созданы
    init_report_structure()

    # Формируем имя файла и вкладки строго по текущей дате
    today_str = datetime.now().strftime("%Y-%m-%d")
    file_path = os.path.join(DAILY_DIR, f"{today_str}.xlsx")
    sheet_title = today_str

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A4"

    # Стилизация (полностью соответствует твоему мастер-отчету)
    font_title = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    font_regular = Font(name="Segoe UI", size=9, color="2C3E50")
    font_bold = Font(name="Segoe UI", size=9, bold=True, color="2C3E50")
    font_link = Font(name="Segoe UI", size=9, color="1A5276", underline="single", bold=True)

    fill_main_banner = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    fill_zebra_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    fill_cat1 = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
    fill_cat2 = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
    fill_cat3 = PatternFill(start_color="FCF3CF", fill_type="solid")
    fill_cat4 = PatternFill(start_color="FADBD8", fill_type="solid")

    border_light = Border(left=Side(style="thin", color="E5E7E9"), right=Side(style="thin", color="E5E7E9"),
                          top=Side(style="thin", color="E5E7E9"), bottom=Side(style="thin", color="E5E7E9"))

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Главный баннер сверху
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"ЕЖЕДНЕВНЫЙ ЖУРНАЛ КОНТРОЛЯ КАЧЕСТВА ЗВОНКОВ — {today_str}"
    title_cell.font = font_title
    title_cell.fill = fill_main_banner
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 32

    headers = [
        "ID Звонка", "Дата", "Время", "Телефон клиента", "Длительность", "Итого балл (/43)", 
        "Внутр. номер", "Направление", "Запись", 
        "Резюме диалога", "Детальный отчёт", "Запись звонка"
    ]

    header_row = 3
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        cell.alignment = align_center
    ws.row_dimensions[header_row].height = 28

    current_row = header_row + 1
    added_rows_count = 0

    for item_tuple in calls_data:
        # Распаковка структуры из БД (словари или кортежи)
        if isinstance(item_tuple, dict):
            start_time_str = item_tuple.get("start_time", "")
            call_id = item_tuple.get("call_id", "")
            duration = item_tuple.get("duration", 0)
            phone = item_tuple.get("customer_phone", "")
            admin_name_from_db = item_tuple.get("admin_name", "Не определен")
            analysis_text = item_tuple.get("analysis_text", "")
            record_url = item_tuple.get("record_url", "")
            direction = item_tuple.get("direction", "in")
            db_report_url = item_tuple.get("report_url", "")
        else:
            start_time_str = item_tuple[0]
            call_id = item_tuple[1]
            duration = item_tuple[2]
            phone = item_tuple[3]
            admin_name_from_db = item_tuple[4]
            # индекс 5 — это clinic_branch, пропускаем его
            analysis_text = item_tuple[6]
            record_url = item_tuple[7]
            direction = item_tuple[8] if len(item_tuple) > 8 else "in"
            db_report_url = item_tuple[9] if len(item_tuple) > 9 else ""

        # Фильтрация по дате: Пропускаем звонки, сделанные НЕ сегодня
        if not start_time_str:
            continue
        try:
            call_date_part = start_time_str.split(" ")[0]
            if call_date_part != today_str:
                continue  # Это звонок за другой день, скипаем
        except Exception:
            continue

        total_score = 0
        appointment_made = False
        dialog_overview = ""

        if analysis_text:
            try:
                parsed = json.loads(analysis_text) if isinstance(analysis_text, str) else analysis_text
                total_score = int(parsed.get("total_score", 0))
                appointment_made = parsed.get("appointment_made", False)
                dialog_overview = parsed.get("dialog_overview", "")
            except:
                pass

        # === ЖЕСТКАЯ ПРОВЕРКА: Если резюме диалога пустое — вообще не пишем строку ===
        if not dialog_overview or str(dialog_overview).strip() in ["", "null", "None", "-"]:
            continue

        ws.row_dimensions[current_row].height = 28
        row_base_fill = fill_zebra_light if added_rows_count % 2 == 0 else None

        call_date = call_time = "-"
        try:
            dt = datetime.strptime(start_time_str.split(".")[0], "%Y-%m-%d %H:%M:%S")
            call_date = dt.strftime("%d.%m.%Y")
            call_time = dt.strftime("%H:%M")
        except:
            pass

        direction_readable = "Входящий" if direction == "in" else "Исходящий"

        row_data = [
            str(call_id), call_date, call_time, str(phone or "-"), int(duration or 0), total_score,
            str(admin_name_from_db or "Не определен"), direction_readable, "Да" if appointment_made else "Нет"
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = font_regular
            cell.border = border_light
            if row_base_fill: cell.fill = row_base_fill
            cell.alignment = align_center if col_idx not in [4, 7] else align_left

        # Цветовая индикация баллов
        score_cell = ws.cell(row=current_row, column=6)
        score_cell.font = font_bold
        if total_score >= 37: score_cell.fill = fill_cat1
        elif total_score >= 28: score_cell.fill = fill_cat2
        elif total_score >= 18: score_cell.fill = fill_cat3
        else: score_cell.fill = fill_cat4

        # Подсветка записи
        booking_cell = ws.cell(row=current_row, column=9)
        if appointment_made:
            booking_cell.fill = fill_cat1
            booking_cell.font = font_bold

        # Описание диалога
        overview_cell = ws.cell(row=current_row, column=10, value=dialog_overview)
        overview_cell.font = font_regular
        overview_cell.border = border_light
        overview_cell.alignment = align_left
        if row_base_fill: overview_cell.fill = row_base_fill

        # === УМНАЯ ССЫЛКА НА ДЕТАЛЬНЫЙ ОТЧЕТ (Колонка K / 11) ===
        # Ищем сначала в оперативной карте синклера, затем в историческом поле БД
        real_public_url = public_urls_map.get(str(call_id)) or db_report_url
        report_cell = ws.cell(row=current_row, column=11)
        
        if real_public_url and str(real_public_url).strip() not in ["", "None", "null", "-"]:
            report_cell.value = "Открыть отчёт"
            report_cell.hyperlink = real_public_url
            report_cell.font = font_link
        else:
            # Относительная локальная ссылка (выход из reports/daily/ в reports/individual/)
            local_rel_path = f"../individual/report_{call_id}.xlsx"
            report_cell.value = "Открыть отчёт (Локально)"
            report_cell.hyperlink = local_rel_path
            report_cell.font = font_link
            
        report_cell.alignment = align_center
        report_cell.border = border_light
        if row_base_fill: report_cell.fill = row_base_fill

        # Ссылка на запись звонка
        link_cell = ws.cell(row=current_row, column=12)
        if record_url:
            link_cell.value = "Слушать запись"
            link_cell.hyperlink = record_url
            link_cell.font = font_link
        else:
            link_cell.value = "-"
            link_cell.font = font_regular
        link_cell.alignment = align_center
        link_cell.border = border_light
        if row_base_fill: link_cell.fill = row_base_fill

        current_row += 1
        added_rows_count += 1

    # Ширина колонок
    column_widths = {
        'A': 14, 'B': 13, 'C': 10, 'D': 18, 'E': 14, 'F': 17, 'G': 28, 'H': 14, 'I': 11,
        'J': 45, 'K': 20, 'L': 18
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(file_path)
    return file_path


# =========================================================================
# ОТДЕЛЬНЫЙ ФАЙЛ СТАТИСТИКИ ПО ВНУТРЕННИМ НОМЕРАМ (по образцу присланной формы)
# =========================================================================

STATS_GROUPS = [
    {
        "title": "ВСЕГО",
        "numbers": {"101", "100", "107", "102", "110", "114", "109", "120", "121", "122", "106", "131", "130", "117"},
        "time_range": None,
        "metrics": ["total", "in", "out", "booked", "lost", "minutes"],
    },
    {
        "title": "ЗВОНКИ админов и операторов",
        "numbers": {"101", "100", "120", "121", "130", "131"},
        "time_range": ("09:01", "21:59"),
        "metrics": ["total", "in", "out", "booked", "lost", "conversion"],
    },
    {
        "title": "ЗВОНКИ ночь Бирюлево",
        "numbers": {"101", "100"},
        "time_range": ("22:00", "09:00"),
        "metrics": ["total", "in", "out", "booked", "lost", "conversion"],
    },
    {
        "title": "ЗВОНКИ админов Бирюлево",
        "numbers": {"101", "100"},
        "time_range": ("09:01", "21:59"),
        "metrics": ["total", "in", "out", "booked", "lost", "conversion"],
    },
    {
        "title": "ЗВОНКИ админов Марьино",
        "numbers": {"120", "121"},
        "time_range": ("09:01", "21:59"),
        "metrics": ["total", "in", "out", "booked", "lost", "conversion"],
    },
    {
        "title": "Коломенское",
        "numbers": {"130", "131"},
        "time_range": ("09:01", "21:59"),
        "metrics": ["total", "in", "out", "booked", "lost", "conversion"],
    },
    {
        "title": "Стационар",
        "numbers": {"114"},
        "time_range": None,
        "metrics": ["total", "in", "out", "lost"],
    },
    {
        "title": "Оператор Раупова Индира",
        "numbers": {"109"},
        "time_range": None,
        "metrics": ["total", "in", "out", "booked", "lost", "conversion"],
    },
    {
        "title": "Оператор Татьяна Изосимова",
        "numbers": {"106"},
        "time_range": None,
        "metrics": ["total", "in", "out", "booked", "lost", "conversion"],
    },
]

METRIC_LABELS = {
    "total": "Всего звонков",
    "in": "Входящие",
    "out": "Исходящие",
    "booked": "Запись",
    "lost": "Потерянные",
    "minutes": "Кол-во минут",
    "conversion": "Конверсия, %",
}


def _in_time_range(call_time, time_range):
    """Проверяет, попадает ли время звонка в заданное окно (с поддержкой ночного окна через полночь)"""
    if not time_range:
        return True
    start_t = datetime.strptime(time_range[0], "%H:%M").time()
    end_t = datetime.strptime(time_range[1], "%H:%M").time()
    if start_t <= end_t:
        return start_t <= call_time <= end_t
    return call_time >= start_t or call_time <= end_t


def _group_events_by_date(all_events: list) -> dict:
    """Группирует события звонков по дате (YYYY-MM-DD, вытащенной из start_time)"""
    events_by_date = {}
    for row in all_events:
        padded = list(row) + [None] * (7 - len(row))
        start_time = padded[1]
        date_key = str(start_time).split(" ")[0].split("T")[0] if start_time else "Не определена"
        events_by_date.setdefault(date_key, []).append(row)
    return events_by_date


def _compute_group_values(events_for_date: list, group: dict) -> dict:
    """Считает метрики одной группы по событиям одной даты"""
    numbers = group["numbers"]
    time_range = group["time_range"]

    total = incoming = outgoing = booked = lost = 0
    minutes_sum = 0.0

    for row in events_for_date:
        padded = list(row) + [None] * (7 - len(row))
        _call_id, start_time, direction, internal_number, duration, is_lost, appointment_made = padded[:7]

        if internal_number not in numbers:
            continue
        try:
            dt = datetime.strptime(str(start_time).split(".")[0], "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
        if not _in_time_range(dt.time(), time_range):
            continue

        total += 1
        if direction == "in":
            incoming += 1
        else:
            outgoing += 1
        if is_lost:
            lost += 1
        if appointment_made:
            booked += 1
        minutes_sum += (duration or 0) / 60

    return {
        "total": total,
        "in": incoming,
        "out": outgoing,
        "booked": booked,
        "lost": lost,
        "minutes": round(minutes_sum, 1),
        "conversion": round((booked / incoming * 100), 1) if incoming else 0,
    }


def update_statistics_report(all_events: list, groups: list = None) -> str:
    """
    Строит ЕДИНЫЙ накопительный файл статистики по внутренним номерам — по одной
    строке на каждую дату, полностью пересобирается при каждом вызове (как
    update_master_report). Лежит на уровне мастер-отчёта, без подпапки.

    all_events — ВСЕ события звонков за всё время из БД (call_id, start_time,
                 direction, internal_number, duration, is_lost, appointment_made).
    """
    if groups is None:
        groups = STATS_GROUPS

    init_report_structure()
    file_path = os.path.join(BASE_REPORTS_DIR, "statistics_report.xlsx")

    # Группируем события по дате (YYYY-MM-DD, вытащенной из start_time)
    events_by_date = _group_events_by_date(all_events)

    dates_sorted = sorted(events_by_date.keys(), reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Статистика"

    font_group_title = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    font_regular = Font(name="Segoe UI", size=10, color="2C3E50")
    font_date = Font(name="Segoe UI", size=10, bold=True, color="2C3E50")

    fill_group = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    fill_header = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    fill_date_col = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_light = Border(left=Side(style="thin", color="D5D8DC"), right=Side(style="thin", color="D5D8DC"),
                          top=Side(style="thin", color="D5D8DC"), bottom=Side(style="thin", color="D5D8DC"))

    # === Заголовки: строка 1 — группы (объединённые), строка 2 — метрики ===
    a2 = ws.cell(row=2, column=1, value="Дата")
    a2.font = font_header
    a2.fill = fill_header
    a2.alignment = align_center
    a2.border = border_light

    col = 2
    for group in groups:
        start_col = col
        for metric in group["metrics"]:
            c = ws.cell(row=2, column=col, value=METRIC_LABELS.get(metric, metric))
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_center
            c.border = border_light
            col += 1
        end_col = col - 1

        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        gc = ws.cell(row=1, column=start_col, value=group["title"])
        gc.font = font_group_title
        gc.fill = fill_group
        gc.alignment = align_center
        for c_idx in range(start_col, end_col + 1):
            ws.cell(row=1, column=c_idx).fill = fill_group
            ws.cell(row=1, column=c_idx).border = border_light

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "B3"

    # === Данные: по одной строке на каждую дату, от новых к старым ===
    data_row = 3
    for date_key in dates_sorted:
        events_for_date = events_by_date[date_key]
        ws.row_dimensions[data_row].height = 20

        d_cell = ws.cell(row=data_row, column=1, value=date_key)
        d_cell.font = font_date
        d_cell.fill = fill_date_col
        d_cell.alignment = align_center
        d_cell.border = border_light

        col = 2
        for group in groups:
            values = _compute_group_values(events_for_date, group)
            for metric in group["metrics"]:
                c = ws.cell(row=data_row, column=col, value=values[metric])
                c.font = font_regular
                c.alignment = align_center
                c.border = border_light
                col += 1

        data_row += 1

    last_col = col if dates_sorted else 2
    ws.column_dimensions['A'].width = 14
    for c_idx in range(2, last_col):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12

    wb.save(file_path)
    return file_path


# =========================================================================
# ДАШБОРД — сводная витрина с графиками и ссылками на остальные отчёты.
# Лежит на уровне мастер-отчёта (reports/dashboard.xlsx), пересобирается
# полностью при каждом вызове, как master/statistics.
# =========================================================================

DASHBOARD_LOOKBACK_DAYS = 14  # сколько последних дней показываем на графиках трендов


def _parse_total_score(analysis_text: str):
    """Достаёт total_score из JSON анализа звонка, если получится"""
    try:
        parsed = json.loads(analysis_text) if analysis_text else {}
        score = parsed.get("total_score")
        return float(score) if score is not None else None
    except Exception:
        return None


def create_dashboard(all_events: list, success_calls: list, links: dict = None) -> str:
    """
    Строит дашборд reports/dashboard.xlsx: KPI-плашки, графики и ссылки на остальные
    отчёты.

    all_events    — все события звонков из get_all_call_events()
                     (call_id, start_time, direction, internal_number, duration, is_lost, appointment_made)
    success_calls — успешные звонки из get_success_calls_for_master()
                     (start_time, call_id, duration, phone, admin_name, clinic_branch,
                      analysis_text, record_url, direction, report_url)
    links         — необязательный dict {название: URL/путь} для блока ссылок,
                     например {"Мастер-отчёт": "master_calls_report.xlsx", ...}
    """
    init_report_structure()
    file_path = os.path.join(BASE_REPORTS_DIR, "dashboard.xlsx")

    if links is None:
        links = {
            "📊 Мастер-отчёт (все звонки)": "master_calls_report.xlsx",
            "📅 Ежедневный отчёт (сегодня)": f"daily/{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            "📈 Статистика по внутренним номерам": "statistics_report.xlsx",
        }

    font_h1 = Font(name="Segoe UI", size=16, bold=True, color="2C3E50")
    font_h2 = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_kpi_value = Font(name="Segoe UI", size=20, bold=True, color="FFFFFF")
    font_kpi_label = Font(name="Segoe UI", size=9, color="FFFFFF")
    font_link = Font(name="Segoe UI", size=10, color="2E75B6", underline="single")
    font_muted = Font(name="Segoe UI", size=9, italic=True, color="7F8C8D")

    fill_kpi_1 = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    fill_kpi_2 = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    fill_kpi_3 = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
    fill_kpi_4 = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    fill_section = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дашборд"
    ws.sheet_view.showGridLines = False

    # === Заголовок ===
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "📋 Дашборд по звонкам клиники"
    title_cell.font = font_h1
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Обновлено: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = font_muted

    # === Блок ссылок на остальные отчёты (строка 3) ===
    link_col = 1
    for label, target in links.items():
        c = ws.cell(row=3, column=link_col, value=label)
        c.font = font_link
        c.hyperlink = target
        link_col += 2
    ws.row_dimensions[3].height = 18

    # === Считаем данные ===
    events_by_date = _group_events_by_date(all_events)
    all_numbers = STATS_GROUPS[0]["numbers"]  # набор "внешних" номеров (группа "ВСЕГО")

    today = datetime.now().date()
    dates_window = [(today - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(DASHBOARD_LOOKBACK_DAYS - 1, -1, -1)]

    trend_rows = []  # (дата, всего, входящие, исходящие, потерянные, запись, конверсия%)
    for date_key in dates_window:
        events_for_date = events_by_date.get(date_key, [])
        vals = _compute_group_values(events_for_date, {"numbers": all_numbers, "time_range": None})
        trend_rows.append((date_key, vals["total"], vals["in"], vals["out"], vals["lost"], vals["booked"], vals["conversion"]))

    total_calls_period = sum(r[1] for r in trend_rows)
    total_lost_period = sum(r[4] for r in trend_rows)
    total_booked_period = sum(r[5] for r in trend_rows)
    total_in_period = sum(r[2] for r in trend_rows)
    avg_conversion = round((total_booked_period / total_in_period * 100), 1) if total_in_period else 0

    # Звонки по группам за сегодня (для столбчатой диаграммы по линиям)
    events_today = events_by_date.get(today.strftime("%Y-%m-%d"), [])
    group_bar_rows = []
    for group in STATS_GROUPS[1:]:  # без "ВСЕГО" — она агрегирующая
        vals = _compute_group_values(events_today, group)
        group_bar_rows.append((group["title"], vals["total"]))

    # Средний балл оценки по линиям (из success_calls)
    scores_by_line = {}
    for row in success_calls:
        padded = list(row) + [None] * (10 - len(row))
        admin_name, analysis_text = padded[4], padded[6]
        score = _parse_total_score(analysis_text)
        if score is None or not admin_name:
            continue
        scores_by_line.setdefault(admin_name, []).append(score)

    avg_scores = sorted(
        [(name, round(sum(v) / len(v), 1), len(v)) for name, v in scores_by_line.items()],
        key=lambda x: x[1]
    )

    # === KPI-плашки (строки 5-7) ===
    kpi_row = 5
    kpis = [
        ("Всего звонков", str(total_calls_period), fill_kpi_1),
        ("Конверсия в запись", f"{avg_conversion}%", fill_kpi_2),
        ("Потерянные", str(total_lost_period), fill_kpi_4),
        ("Записей создано", str(total_booked_period), fill_kpi_3),
    ]
    kpi_col = 1
    for label, value, fill in kpis:
        ws.merge_cells(start_row=kpi_row, start_column=kpi_col, end_row=kpi_row, end_column=kpi_col + 1)
        vcell = ws.cell(row=kpi_row, column=kpi_col, value=value)
        vcell.font = font_kpi_value
        vcell.fill = fill
        vcell.alignment = align_center
        ws.cell(row=kpi_row, column=kpi_col + 1).fill = fill

        ws.merge_cells(start_row=kpi_row + 1, start_column=kpi_col, end_row=kpi_row + 1, end_column=kpi_col + 1)
        lcell = ws.cell(row=kpi_row + 1, column=kpi_col, value=label)
        lcell.font = font_kpi_label
        lcell.fill = fill
        lcell.alignment = Alignment(horizontal="center", vertical="bottom")
        ws.cell(row=kpi_row + 1, column=kpi_col + 1).fill = fill

        kpi_col += 2
    ws.row_dimensions[kpi_row].height = 30
    ws.row_dimensions[kpi_row + 1].height = 16
    ws.freeze_panes = "A4"
    ws["A4"] = f"За последние {DASHBOARD_LOOKBACK_DAYS} дней"
    ws["A4"].font = font_muted

    # === Скрытая область данных для графиков (лист "Данные") ===
    data_ws = wb.create_sheet("Данные")
    data_ws.sheet_state = "hidden"

    # -- Тренд по дням --
    data_ws.append(["Дата", "Всего", "Входящие", "Исходящие", "Потерянные", "Запись", "Конверсия %"])
    for row in trend_rows:
        data_ws.append(list(row))
    trend_last_row = 1 + len(trend_rows)

    # -- Звонки по группам сегодня --
    group_start_row = trend_last_row + 3
    data_ws.cell(row=group_start_row, column=1, value="Группа")
    data_ws.cell(row=group_start_row, column=2, value="Звонков")
    for i, (name, total) in enumerate(group_bar_rows, start=1):
        data_ws.cell(row=group_start_row + i, column=1, value=name)
        data_ws.cell(row=group_start_row + i, column=2, value=total)
    group_last_row = group_start_row + len(group_bar_rows)

    # -- Направление (сегодня, все номера) --
    direction_start_row = group_last_row + 3
    total_in_today = sum(1 for r in events_today if (list(r) + [None] * 7)[2] == "in")
    total_out_today = sum(1 for r in events_today if (list(r) + [None] * 7)[2] != "in")
    data_ws.cell(row=direction_start_row, column=1, value="Направление")
    data_ws.cell(row=direction_start_row, column=2, value="Звонков")
    data_ws.cell(row=direction_start_row + 1, column=1, value="Входящие")
    data_ws.cell(row=direction_start_row + 1, column=2, value=total_in_today)
    data_ws.cell(row=direction_start_row + 2, column=1, value="Исходящие")
    data_ws.cell(row=direction_start_row + 2, column=2, value=total_out_today)
    direction_last_row = direction_start_row + 2

    # -- Средний балл по линиям (снизу вверх — худшие сверху для наглядности) --
    scores_start_row = direction_last_row + 3
    data_ws.cell(row=scores_start_row, column=1, value="Линия")
    data_ws.cell(row=scores_start_row, column=2, value="Средний балл")
    for i, (name, avg, cnt) in enumerate(avg_scores, start=1):
        data_ws.cell(row=scores_start_row + i, column=1, value=f"{name} ({cnt})")
        data_ws.cell(row=scores_start_row + i, column=2, value=avg)
    scores_last_row = scores_start_row + len(avg_scores)

    # === Графики на главном листе ===
    chart_row = 8

    # 1) Линейный график: конверсия и всего звонков по дням
    line_chart = LineChart()
    line_chart.title = "Динамика звонков и конверсии по дням"
    line_chart.style = 10
    line_chart.y_axis.title = "Звонков"
    line_chart.x_axis.title = "Дата"
    cats = Reference(data_ws, min_col=1, min_row=2, max_row=trend_last_row)
    data_total = Reference(data_ws, min_col=2, min_row=1, max_row=trend_last_row)
    data_conv = Reference(data_ws, min_col=7, min_row=1, max_row=trend_last_row)
    line_chart.add_data(data_total, titles_from_data=True)
    line_chart.add_data(data_conv, titles_from_data=True)
    line_chart.set_categories(cats)
    line_chart.width = 24
    line_chart.height = 10
    ws.add_chart(line_chart, f"A{chart_row}")

    # 2) Столбчатая диаграмма: звонки по группам/линиям сегодня
    if group_bar_rows:
        bar_chart = BarChart()
        bar_chart.title = "Звонки по линиям сегодня"
        bar_chart.style = 12
        bar_chart.y_axis.title = "Звонков"
        cats2 = Reference(data_ws, min_col=1, min_row=group_start_row + 1, max_row=group_last_row)
        data2 = Reference(data_ws, min_col=2, min_row=group_start_row, max_row=group_last_row)
        bar_chart.add_data(data2, titles_from_data=True)
        bar_chart.set_categories(cats2)
        bar_chart.width = 24
        bar_chart.height = 10
        ws.add_chart(bar_chart, f"J{chart_row}")

    chart_row_2 = chart_row + 21

    # 3) Круговая диаграмма: направление звонков сегодня
    pie_chart = PieChart()
    pie_chart.title = "Входящие / исходящие сегодня"
    cats3 = Reference(data_ws, min_col=1, min_row=direction_start_row + 1, max_row=direction_last_row)
    data3 = Reference(data_ws, min_col=2, min_row=direction_start_row, max_row=direction_last_row)
    pie_chart.add_data(data3, titles_from_data=True)
    pie_chart.set_categories(cats3)
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.width = 12
    pie_chart.height = 10
    ws.add_chart(pie_chart, f"A{chart_row_2}")

    # 4) Столбчатая диаграмма: средний балл оценки по линиям
    if avg_scores:
        score_chart = BarChart()
        score_chart.type = "bar"  # горизонтальные столбцы — удобнее читать названия линий
        score_chart.title = "Средний балл оценки по линиям (/43)"
        score_chart.style = 11
        cats4 = Reference(data_ws, min_col=1, min_row=scores_start_row + 1, max_row=scores_last_row)
        data4 = Reference(data_ws, min_col=2, min_row=scores_start_row, max_row=scores_last_row)
        score_chart.add_data(data4, titles_from_data=True)
        score_chart.set_categories(cats4)
        score_chart.width = 24
        score_chart.height = 12
        ws.add_chart(score_chart, f"J{chart_row_2}")

    # === Ширины колонок на главном листе ===
    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 14

    wb.save(file_path)
    return file_path
